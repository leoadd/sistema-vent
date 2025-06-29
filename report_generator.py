import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime, timedelta #timedelta añadido
import os # Para manejar rutas de archivos

# --- Estilos y Helpers ---
BASE_REPORTS_DIR = "generated_reports" # Carpeta para guardar reportes

def _ensure_reports_dir():
    if not os.path.exists(BASE_REPORTS_DIR):
        os.makedirs(BASE_REPORTS_DIR)

def _get_report_filename(base_name, extension):
    _ensure_reports_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(BASE_REPORTS_DIR, f"{base_name}_{timestamp}.{extension}")

# Estilos para Excel
font_bold = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Azul oscuro
alignment_center = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Estilos para PDF
pdf_styles = getSampleStyleSheet()
pdf_title_style = pdf_styles['h1']
pdf_header_style = ParagraphStyle('TableHeader', parent=pdf_styles['Normal'], fontName='Helvetica-Bold', alignment=1) # Alignment 1 = CENTER
pdf_body_style = pdf_styles['Normal']
pdf_body_style.alignment = 0 # Izquierda
pdf_body_right_style = ParagraphStyle('BodyRight', parent=pdf_styles['Normal'], alignment=2) # Derecha


# --- Generador de Reporte de Ventas por Período ---
def generar_excel_ventas_periodo(datos_reporte, fecha_inicio, fecha_fin):
    filename = _get_report_filename("ReporteVentasPeriodo", "xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Ventas"

    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de Ventas del {fecha_inicio} al {fecha_fin}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = alignment_center

    # Resumen General
    resumen = datos_reporte.get("resumen", {})
    ws['A3'] = "Resumen General"; ws['A3'].font = font_bold; ws['A3'].fill = header_fill
    ws.merge_cells('A3:B3')
    ws['A4'] = "Total Transacciones:"; ws['B4'] = resumen.get('numero_transacciones', 0)
    ws['A5'] = "Total Ventas ($):"; ws['B5'] = resumen.get('total_ventas_periodo', 0); ws['B5'].number_format = '"$"#,##0.00'
    ws['A6'] = "Venta Promedio ($):"; ws['B6'] = resumen.get('venta_promedio', 0); ws['B6'].number_format = '"$"#,##0.00'

    # Desglose por Tipo de Pago
    ws['D3'] = "Ventas por Tipo de Pago"; ws['D3'].font = font_bold; ws['D3'].fill = header_fill
    ws.merge_cells('D3:F3')
    headers_tipo_pago = ["Tipo de Pago", "Transacciones", "Total ($)"]
    for col_num, header in enumerate(headers_tipo_pago, 4): # Empezar en columna D
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = font_bold; cell.fill = header_fill; cell.alignment = alignment_center; cell.border = thin_border

    row_idx = 5
    for tipo_pago_data in datos_reporte.get("desglose_tipo_pago", []):
        ws.cell(row=row_idx, column=4, value=tipo_pago_data.get('tipo_pago','N/A')).border = thin_border
        ws.cell(row=row_idx, column=5, value=tipo_pago_data.get('transacciones_por_tipo',0)).border = thin_border
        cell_total = ws.cell(row=row_idx, column=6, value=tipo_pago_data.get('total_por_tipo',0))
        cell_total.number_format = '"$"#,##0.00'; cell_total.border = thin_border
        row_idx += 1

    # Ajustar ancho de columnas
    for col_letter in ['A', 'B', 'D', 'E', 'F']: ws.column_dimensions[col_letter].width = 20

    wb.save(filename)
    return filename

def generar_pdf_ventas_periodo(datos_reporte, fecha_inicio, fecha_fin):
    filename = _get_report_filename("ReporteVentasPeriodo", "pdf")
    doc = SimpleDocTemplate(filename, pagesize=letter)
    elements = []

    elements.append(Paragraph(f"Reporte de Ventas: {fecha_inicio} a {fecha_fin}", pdf_title_style))
    elements.append(Spacer(1, 0.2*inch))

    resumen = datos_reporte.get("resumen", {})
    elements.append(Paragraph("<b>Resumen General:</b>", pdf_styles['h2']))
    elements.append(Paragraph(f"Total Transacciones: {resumen.get('numero_transacciones',0)}", pdf_body_style))
    elements.append(Paragraph(f"Total Ventas: ${resumen.get('total_ventas_periodo',0):.2f}", pdf_body_style))
    elements.append(Paragraph(f"Venta Promedio: ${resumen.get('venta_promedio',0):.2f}", pdf_body_style))
    elements.append(Spacer(1, 0.2*inch))

    elements.append(Paragraph("<b>Ventas por Tipo de Pago:</b>", pdf_styles['h2']))
    data_tipo_pago = [["Tipo de Pago", "Transacciones", "Total ($)"]]
    for item in datos_reporte.get("desglose_tipo_pago", []):
        data_tipo_pago.append([item.get('tipo_pago','N/A'), str(item.get('transacciones_por_tipo',0)), f"${item.get('total_por_tipo',0):.2f}"])

    table_tipo_pago = Table(data_tipo_pago, colWidths=[2*inch, 1.5*inch, 1.5*inch])
    table_tipo_pago.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12), ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(table_tipo_pago)

    doc.build(elements)
    return filename

# --- Generador de Reporte de Productos Más Vendidos ---
def generar_excel_productos_mas_vendidos(datos_reporte, fecha_inicio, fecha_fin, top_n):
    filename = _get_report_filename(f"Top{top_n}Productos", "xlsx")
    wb = openpyxl.Workbook()

    # Hoja por Cantidad
    ws_qty = wb.active
    ws_qty.title = f"Top {top_n} por Cantidad"
    ws_qty.merge_cells('A1:D1'); ws_qty['A1'] = f"Top {top_n} Productos Más Vendidos por Cantidad ({fecha_inicio} a {fecha_fin})"; ws_qty['A1'].font = Font(size=14, bold=True)
    headers_qty = ["Código Barras", "Nombre Producto", "Total Cantidad Vendida"]
    for col, header in enumerate(headers_qty, 1): ws_qty.cell(row=3, column=col, value=header).font = font_bold; ws_qty.cell(row=3, column=col, value=header).fill = header_fill

    for row_idx, item in enumerate(datos_reporte.get("top_por_cantidad", []), 4):
        ws_qty.cell(row=row_idx, column=1, value=item.get('codigo_barras','N/A'))
        ws_qty.cell(row=row_idx, column=2, value=item.get('nombre_producto','N/A'))
        ws_qty.cell(row=row_idx, column=3, value=item.get('total_cantidad_vendida',0))
    ws_qty.column_dimensions['B'].width = 40; ws_qty.column_dimensions['A'].width = 20; ws_qty.column_dimensions['C'].width = 25


    # Hoja por Valor
    ws_val = wb.create_sheet(f"Top {top_n} por Valor")
    ws_val.merge_cells('A1:D1'); ws_val['A1'] = f"Top {top_n} Productos Más Vendidos por Valor Total ({fecha_inicio} a {fecha_fin})"; ws_val['A1'].font = Font(size=14, bold=True)
    headers_val = ["Código Barras", "Nombre Producto", "Total Valor Vendido ($)"]
    for col, header in enumerate(headers_val, 1): ws_val.cell(row=3, column=col, value=header).font = font_bold; ws_val.cell(row=3, column=col, value=header).fill = header_fill

    for row_idx, item in enumerate(datos_reporte.get("top_por_valor", []), 4):
        ws_val.cell(row=row_idx, column=1, value=item.get('codigo_barras','N/A'))
        ws_val.cell(row=row_idx, column=2, value=item.get('nombre_producto','N/A'))
        cell_val = ws_val.cell(row=row_idx, column=3, value=item.get('total_valor_vendido',0))
        cell_val.number_format = '"$"#,##0.00'
    ws_val.column_dimensions['B'].width = 40; ws_val.column_dimensions['A'].width = 20; ws_val.column_dimensions['C'].width = 25

    wb.save(filename)
    return filename

def generar_pdf_productos_mas_vendidos(datos_reporte, fecha_inicio, fecha_fin, top_n):
    filename = _get_report_filename(f"Top{top_n}Productos", "pdf")
    doc = SimpleDocTemplate(filename, pagesize=letter)
    elements = []
    elements.append(Paragraph(f"Top {top_n} Productos Más Vendidos ({fecha_inicio} a {fecha_fin})", pdf_title_style))

    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(f"<b>Por Cantidad Vendida:</b>", pdf_styles['h2']))
    data_qty = [["Código", "Producto", "Cantidad Total"]]
    for item in datos_reporte.get("top_por_cantidad",[]):
        data_qty.append([item.get('codigo_barras','-'), Paragraph(item.get('nombre_producto','-'),pdf_body_style), str(item.get('total_cantidad_vendida',0))])
    table_qty = Table(data_qty, colWidths=[1.5*inch, 3.5*inch, 1.5*inch])
    table_qty.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'LEFT'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('GRID',(0,0),(-1,-1),1,colors.black), ('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    elements.append(table_qty)

    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"<b>Por Valor Total Vendido:</b>", pdf_styles['h2']))
    data_val = [["Código", "Producto", "Valor Total ($)"]]
    for item in datos_reporte.get("top_por_valor",[]):
        data_val.append([item.get('codigo_barras','-'), Paragraph(item.get('nombre_producto','-'),pdf_body_style), f"${item.get('total_valor_vendido',0):.2f}"])
    table_val = Table(data_val, colWidths=[1.5*inch, 3.5*inch, 1.5*inch])
    table_val.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'LEFT'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('GRID',(0,0),(-1,-1),1,colors.black), ('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    elements.append(table_val)

    doc.build(elements)
    return filename

# --- Generador de Reporte de Ventas por Usuario ---
def generar_excel_ventas_por_usuario(datos_reporte, fecha_inicio, fecha_fin):
    filename = _get_report_filename("VentasPorUsuario", "xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas por Usuario"
    ws.merge_cells('A1:C1'); ws['A1'] = f"Reporte de Ventas por Usuario ({fecha_inicio} a {fecha_fin})"; ws['A1'].font = Font(size=14, bold=True)
    headers = ["Nombre de Usuario", "Número de Ventas", "Total Vendido ($)"]
    for col, header in enumerate(headers, 1): ws.cell(row=3, column=col, value=header).font = font_bold; ws.cell(row=3,column=col).fill = header_fill

    for row_idx, item in enumerate(datos_reporte, 4): # datos_reporte es una lista aquí
        ws.cell(row=row_idx, column=1, value=item.get('nombre_usuario','N/A'))
        ws.cell(row=row_idx, column=2, value=item.get('numero_ventas',0))
        cell_total = ws.cell(row=row_idx, column=3, value=item.get('total_vendido_por_usuario',0))
        cell_total.number_format = '"$"#,##0.00'
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 20
    wb.save(filename)
    return filename

def generar_pdf_ventas_por_usuario(datos_reporte, fecha_inicio, fecha_fin):
    filename = _get_report_filename("VentasPorUsuario", "pdf")
    doc = SimpleDocTemplate(filename, pagesize=letter)
    elements = []
    elements.append(Paragraph(f"Reporte de Ventas por Usuario ({fecha_inicio} a {fecha_fin})", pdf_title_style))
    elements.append(Spacer(1, 0.2*inch))

    data = [["Usuario", "Nº Ventas", "Total Vendido ($)"]]
    for item in datos_reporte: # datos_reporte es una lista
        data.append([Paragraph(item.get('nombre_usuario','-'), pdf_body_style), str(item.get('numero_ventas',0)), f"${item.get('total_vendido_por_usuario',0):.2f}"])

    table = Table(data, colWidths=[3*inch, 1.5*inch, 2*inch])
    table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'LEFT'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('GRID',(0,0),(-1,-1),1,colors.black),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    elements.append(table)
    doc.build(elements)
    return filename

if __name__ == '__main__':
    print(f"Módulo generador de reportes. Los reportes se guardarán en '{BASE_REPORTS_DIR}/'")
    # Crear datos de ejemplo para probar las funciones (simulando salida de los controllers)
    _ensure_reports_dir() # Asegurarse de que el directorio existe

    hoy = datetime.now().strftime("%Y-%m-%d")
    hace_7_dias = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")

    # Ejemplo para Reporte de Ventas por Período
    datos_ventas_periodo_ej = {
        "resumen": {"numero_transacciones": 150, "total_ventas_periodo": 7580.50, "venta_promedio": 50.54},
        "desglose_tipo_pago": [
            {"tipo_pago": "efectivo", "transacciones_por_tipo": 100, "total_por_tipo": 5000.00},
            {"tipo_pago": "tarjeta_debito", "transacciones_por_tipo": 50, "total_por_tipo": 2580.50}
        ]
    }
    print(f"Generando Excel Ventas Período: {generar_excel_ventas_periodo(datos_ventas_periodo_ej, hace_7_dias, hoy)}")
    print(f"Generando PDF Ventas Período: {generar_pdf_ventas_periodo(datos_ventas_periodo_ej, hace_7_dias, hoy)}")

    # Ejemplo para Productos Más Vendidos
    datos_top_productos_ej = {
        "top_por_cantidad": [
            {"codigo_barras": "P001", "nombre_producto": "Producto Estrella A", "total_cantidad_vendida": 120},
            {"codigo_barras": "P005", "nombre_producto": "Producto Popular B", "total_cantidad_vendida": 95},
        ],
        "top_por_valor": [
            {"codigo_barras": "P001", "nombre_producto": "Producto Estrella A", "total_valor_vendido": 2400.00},
            {"codigo_barras": "P003", "nombre_producto": "Producto Caro C", "total_valor_vendido": 1500.00},
        ]
    }
    print(f"Generando Excel Top Productos: {generar_excel_productos_mas_vendidos(datos_top_productos_ej, hace_7_dias, hoy, 2)}")
    print(f"Generando PDF Top Productos: {generar_pdf_productos_mas_vendidos(datos_top_productos_ej, hace_7_dias, hoy, 2)}")

    # Ejemplo para Ventas por Usuario
    datos_ventas_usuario_ej = [
        {"nombre_usuario": "vendedor1", "numero_ventas": 80, "total_vendido_por_usuario": 4000.00},
        {"nombre_usuario": "vendedor2", "numero_ventas": 70, "total_vendido_por_usuario": 3580.50},
    ]
    print(f"Generando Excel Ventas por Usuario: {generar_excel_ventas_por_usuario(datos_ventas_usuario_ej, hace_7_dias, hoy)}")
    print(f"Generando PDF Ventas por Usuario: {generar_pdf_ventas_por_usuario(datos_ventas_usuario_ej, hace_7_dias, hoy)}")

    print("Pruebas de generación de reportes completadas.")

# --- Generador de Excel para Productos ---
def generar_excel_productos(lista_productos):
    filename = _get_report_filename("ListadoProductos", "xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Título
    ws.merge_cells('A1:M1') # Ajustar el número de columnas según los headers
    ws['A1'] = "Listado Completo de Productos"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = alignment_center

    # Cabeceras
    headers = [
        "ID", "Código Barras", "Nombre Producto", "Descripción",
        "Categoría ID", "Categoría", "Proveedor ID", "Proveedor",
        "Precio Compra ($)", "Precio Venta ($)", "Precio Mayoreo ($)", "Cant. Mayoreo",
        "Stock Actual", "Stock Mínimo", "Unidad Medida", "Activo",
        "Fecha Creación", "Última Modificación"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = font_bold; cell.fill = header_fill; cell.alignment = alignment_center; cell.border = thin_border

    # Datos de productos
    row_idx = 4
    for prod in lista_productos: # Asumimos que lista_productos es una lista de dicts como los de models.listar_productos
        ws.cell(row=row_idx, column=1, value=prod.get('id')).border = thin_border
        ws.cell(row=row_idx, column=2, value=prod.get('codigo_barras')).border = thin_border
        ws.cell(row=row_idx, column=3, value=prod.get('nombre_producto')).border = thin_border
        ws.cell(row=row_idx, column=4, value=prod.get('descripcion')).border = thin_border
        ws.cell(row=row_idx, column=5, value=prod.get('categoria_id')).border = thin_border
        ws.cell(row=row_idx, column=6, value=prod.get('nombre_categoria')).border = thin_border # Asumiendo que listar_productos hace JOIN
        ws.cell(row=row_idx, column=7, value=prod.get('proveedor_id')).border = thin_border
        ws.cell(row=row_idx, column=8, value=prod.get('nombre_proveedor')).border = thin_border # Asumiendo que listar_productos hace JOIN

        pc = ws.cell(row=row_idx, column=9, value=prod.get('precio_compra', 0))
        pc.number_format = '"$"#,##0.00'; pc.border = thin_border
        pv = ws.cell(row=row_idx, column=10, value=prod.get('precio_venta_menudeo', 0))
        pv.number_format = '"$"#,##0.00'; pv.border = thin_border
        pm = ws.cell(row=row_idx, column=11, value=prod.get('precio_venta_mayoreo')) # Puede ser None
        if pm.value is not None: pm.number_format = '"$"#,##0.00'
        pm.border = thin_border

        ws.cell(row=row_idx, column=12, value=prod.get('cantidad_para_mayoreo')).border = thin_border
        ws.cell(row=row_idx, column=13, value=prod.get('stock_actual', 0)).border = thin_border
        ws.cell(row=row_idx, column=14, value=prod.get('stock_minimo', 0)).border = thin_border
        ws.cell(row=row_idx, column=15, value=prod.get('unidad_medida')).border = thin_border
        ws.cell(row=row_idx, column=16, value="Sí" if prod.get('activo', True) else "No").border = thin_border
        ws.cell(row=row_idx, column=17, value=prod.get('fecha_creacion')).border = thin_border
        ws.cell(row=row_idx, column=18, value=prod.get('fecha_ultima_modificacion')).border = thin_border
        row_idx += 1

    # Ajustar ancho de columnas (ejemplo)
    ws.column_dimensions['C'].width = 35 # Nombre Producto
    ws.column_dimensions['D'].width = 30 # Descripcion
    for i in range(1, len(headers) + 1):
        if ws.column_dimensions[get_column_letter(i)].width < 15 : # Ancho mínimo para otras
             if get_column_letter(i) not in ['C','D']:
                 ws.column_dimensions[get_column_letter(i)].width = 15

    wb.save(filename)
    return filename
